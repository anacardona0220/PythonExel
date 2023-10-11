import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference 
from openpyxl.styles import Font 
import string


archivo_excel = pd.read_excel('./supermarket_sales.xlsx')
# print(archivo_excel[['Gender', 'Product line', 'Total']])

tabla_pivote = archivo_excel.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
# print(tabla_pivote)

tabla_pivote.to_excel('Sales_2021.xlsx', startrow=4, sheet_name='report')

wb = load_workbook('Sales_2021.xlsx')
hoja = wb['report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_rows = wb.active.min_row
max_rows = wb.active.max_row

# print(min_col)
# print(max_col)
# print(min_row)
# print(max_row)

# -------------- Graficar ------------

barChart = BarChart()
data = Reference(hoja, min_col=min_column+1, max_col=max_column, min_row=min_rows, max_row=max_rows)
category = Reference(hoja, min_col=min_column, max_col=min_column, min_row=min_rows+1, max_row=max_rows)

barChart.add_data(data, titles_from_data=True)
barChart.set_categories(category)

hoja.add_chart(barChart, 'B12')
barChart.title = 'Ventas'
barChart.style = 5
# -------------- formulas ------------
# hoja['B8'] = '=SUM(B6:B7)' #Una sola celda

abecedario = list(string.ascii_uppercase)
abecedario_Excel = abecedario[0:max_column]

for i in abecedario_Excel:
    if i != 'A':
        hoja[f'{i}{max_rows+1}'] = f'=SUM({i}{min_rows+1}:{i}{max_rows})'
        hoja[f'{i}{max_rows+1}'].style = 'Currency'
        
hoja[f'{abecedario_Excel[0]}{max_rows+1}'] = 'Total'
        
    
# -------------- Titulos ------------

hoja['A1'] = 'Reporte'
hoja['A2'] = '2023'

hoja['A1'].font = Font('Arial', bold=True, size=22)
hoja['A1'].font = Font('Arial', bold=True, size=16)
    
wb.save('Sales_2021.xlsx')
